from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wdw
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException as te
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime
import openpyxl as XL
import os

TP = 2
timeout = 120
filePath = '1_Raw_Files\\Input'
Path = os.path.join(os.path.dirname(os.path.abspath(__name__)), filePath)
LogPath = os.path.join(os.path.dirname(
    os.path.abspath(__name__)), 'WABMS-Logs')


ChromePath = os.path.join(os.path.dirname(
    os.path.abspath(__name__)), 'chromedriver')

print(ChromePath)
inputFile = []


def getInputFile():
    for Dir, SubDir, Files in os.walk(Path):
        for file in Files:
            if file.endswith('.xlsx'):
                inputFile.append(os.path.join(Path, file))


getInputFile()


inputData = []


def readInputFile(file):
    WB = XL.load_workbook(file[0], data_only=True)
    Sheet = WB[WB.sheetnames[0]]
    Val = int(input('Enter the send count: '))

    for rows in Sheet.iter_rows(min_row=2, max_row=Val, values_only=True):
        dataDoj = {
            'Name': rows[0],
            'Number': rows[1],
            'Message': rows[2],
        }
        inputData.append(dataDoj)


readInputFile(inputFile)


def Drive():
    options = Options()
    # options.add_argument('--start-maximized')
    return wd.Chrome(ChromePath, options=options)


D = Drive()


def ElemX(E):
    time.sleep(TP)
    return D.find_element_by_xpath(E)


def ElemXCheck(e, mess, data):
    time.sleep(TP)
    try:
        wdw(D, timeout).until(ec.visibility_of_element_located(
            (By.XPATH, e)))
        return D.find_element_by_xpath(e)
    except te:
        print(f'{mess} element is not not visible.')
        D.quit()


def Start():
    D.get(f'http://wa.me/{inputData[0]["Number"]}')
    time.sleep(5)
    ElemX('//*[@id="action-button"]').click()
    time.sleep(2)
    ElemX('//*[@id="fallback_block"]/div/div/a').click()
    time.sleep(3)
    time.sleep(1)

    try:
        wdw(D, timeout).until(ec.visibility_of_element_located(
            (By.XPATH, '//*[@id="side"]/header/div[2]/div/span/div[3]/div/span')))
        print('Whats App QR Code Scan Completed.')

    except te:
        print('You did not scan WhatsApp QR Code. Closing now!!!')
        D.quit()


Start()


def StaRt():
    for I, i in enumerate(inputData):

        print(f'i --- {i}')
        print(f'I --- {I}')
        print(i['Number'])
        if I == 0:
            time.sleep(2)
            ElemXCheck(
                '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]', 'Message box text', i).send_keys(i["Message"])

            ElemXCheck('//*[@id="main"]/footer/div[1]/div[3]/button',
                       'Message send button', i).click()
            time.sleep(10)

            with open(os.path.join(os.path.dirname(os.path.abspath(__name__)), 'WABMS-Logs', f'WhatsApp-Bulk-Message-Send-log-{str(datetime.now().date())}.log'), 'a') as file:
                file.write(
                    f"{i['Number']}-{i['Name']}-{i['Message']}-{datetime.now()}\n")

        if I != 0:
            D.get(f'http://wa.me/{i["Number"]}')
            ElemX('//*[@id="action-button"]').click()
            ElemX('//*[@id="fallback_block"]/div/div/a').click()
            time.sleep(2)

            ElemXCheck(
                '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]', 'Message box text', i).send_keys(i["Message"])

            ElemXCheck('//*[@id="main"]/footer/div[1]/div[3]/button',
                       'Message send button', i).click()
            time.sleep(10)

            with open(os.path.join(os.path.dirname(os.path.abspath(__name__)), 'WABMS-Logs', f'WhatsApp-Bulk-Message-Send-log-{str(datetime.now().date())}.log'), 'a') as file:
                file.write(
                    f"{i['Number']}-{i['Name']}-{i['Message']}-{datetime.now()}\n")


StaRt()
